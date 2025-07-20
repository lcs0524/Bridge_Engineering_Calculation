# -*- coding: utf-8 -*-
"""
结果导出器 - 导出计算结果到各种格式
"""

import pandas as pd
import os
from datetime import datetime
import json
import openpyxl


class ResultExporter:
    """结果导出器类"""
    
    def __init__(self):
        """初始化导出器"""
        pass
    
    def export_to_excel(self, results, filename):
        """
        导出结果到Excel文件
        
        参数:
        results: 计算结果字典
        filename: 输出文件名
        """
        try:
            # 创建Excel写入器
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # 1. 输入参数表
                self._write_input_parameters(writer, results)
                
                # 2. 计算结果表
                self._write_calculation_results(writer, results)
                
                # 3. 统计分析表
                self._write_statistics(writer, results)
                
                # 4. 安全评估表
                self._write_safety_assessment(writer, results)
                
                # 5. 计算简图
                self._write_calculation_diagram(writer, results, results.get('input_parameters', {}))
                
                # 6. 修正系数表
                self._write_correction_factors(writer, results)
            
            return True, "Excel导出成功"
            
        except Exception as e:
            return False, f"Excel导出失败：{str(e)}"
    
    def _write_input_parameters(self, writer, results):
        """写入输入参数表"""
        input_params = results['input_parameters']
        
        # 创建输入参数表
        sheet_name = '输入参数'
        writer.sheets[sheet_name] = writer.book.create_sheet(sheet_name)
        
        # 添加项目名称
        row = 1
        writer.sheets[sheet_name].cell(row=row, column=1, value="[ 项目名称 ]")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value=input_params.get('project_name', '高架桥桩基沉降分析项目'))
        row += 2
        
        # 添加项目类型
        writer.sheets[sheet_name].cell(row=row, column=1, value="[ 项目类型 ]")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value="----------------------------------------------------------------------")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value=input_params.get('project_type', '桥梁工程'))
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value="*5种选择其中1种，容许值在计算过程中参考*")
        row += 2
        
        # 计算条件
        writer.sheets[sheet_name].cell(row=row, column=1, value="[ 计算条件 ]")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value="----------------------------------------------------------------------")
        row += 1
        
        # 新建公路信息
        writer.sheets[sheet_name].cell(row=row, column=1, value="[ 新建公路 ]")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value=f"公路类型：{input_params.get('road_level', '一级公路')}")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value=f"车道数量：{input_params.get('lane_count', 4)}")
        row += 2
        
        # 荷载信息
        if 'pile1' in input_params:
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩顶荷载P1：{input_params['pile1']['load']} (KN)")
        if 'pile2' in input_params:
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩顶荷载P2：{input_params['pile2']['load']} (KN)")
        row += 2
        
        # 土层参数
        writer.sheets[sheet_name].cell(row=row, column=1, value="[ 土层参数 ]")
        row += 1
        writer.sheets[sheet_name].cell(row=row, column=1, value="沉降计算修正公式：")
        row += 2
        
        # 土层参数表格
        soil_layers = input_params.get('soil_layers', [])
        if soil_layers:
            # 表头
            writer.sheets[sheet_name].cell(row=row, column=1, value="序号")
            writer.sheets[sheet_name].cell(row=row, column=2, value="土类型")
            writer.sheets[sheet_name].cell(row=row, column=3, value="土层厚")
            writer.sheets[sheet_name].cell(row=row, column=4, value="压缩模量 E")
            writer.sheets[sheet_name].cell(row=row, column=5, value="泊松比 v")
            row += 1
            
            # 土层数据
            for i, layer in enumerate(soil_layers):
                writer.sheets[sheet_name].cell(row=row, column=1, value=f"{i+1}")
                writer.sheets[sheet_name].cell(row=row, column=2, value=layer.get('name', 'N/A'))
                
                # 从depth_range计算土层厚度
                depth_range = layer.get('depth_range', 'N/A')
                thickness = 'N/A'
                if '-' in str(depth_range):
                    try:
                        start, end = map(float, depth_range.split('-'))
                        thickness = end - start
                    except:
                        pass
                
                writer.sheets[sheet_name].cell(row=row, column=3, value=thickness)
                writer.sheets[sheet_name].cell(row=row, column=4, value=layer.get('compression_modulus', 'N/A'))
                writer.sheets[sheet_name].cell(row=row, column=5, value=layer.get('poisson_ratio', 'N/A'))
                row += 1
        
        row += 2
        
        # 桩1参数
        if 'pile1' in input_params:
            writer.sheets[sheet_name].cell(row=row, column=1, value="[ 桩1参数 ]")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩直径：{input_params['pile1']['diameter']} (m)")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩长：{input_params['pile1']['length']} (m)")
            row += 2
        
        # 桩2参数
        if 'pile2' in input_params:
            writer.sheets[sheet_name].cell(row=row, column=1, value="[ 桩2参数 ]")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩直径：{input_params['pile2']['diameter']} (m)")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"桩长：{input_params['pile2']['length']} (m)")
            row += 2
        
        # 被跨越公路参数
        if 'road_params' in input_params:
            road_params = input_params['road_params']
            writer.sheets[sheet_name].cell(row=row, column=1, value="[ 被跨越公路参数 ]")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"路基宽度：{road_params.get('width', 'N/A')} (m)")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"路基与桩1距离：{road_params.get('pile1_distance', 'N/A')} (m)")
            row += 1
            writer.sheets[sheet_name].cell(row=row, column=1, value=f"路基与桩2距离：{road_params.get('pile2_distance', 'N/A')} (m)")
        
        # 调整列宽
        for i in range(1, 6):
            writer.sheets[sheet_name].column_dimensions[chr(64+i)].width = 15
    
    def _write_calculation_results(self, writer, results):
        """写入计算结果表"""
        points = results['points']
        safety_assessment = results.get('safety_assessment', {})
        
        # 表头和说明
        sheet_name = '计算结果'
        title_row = 0
        writer.sheets[sheet_name] = writer.book.create_sheet(sheet_name)
        
        # 添加标题
        writer.sheets[sheet_name].cell(row=title_row+1, column=1, value="计算结果:")
        writer.sheets[sheet_name].cell(row=title_row+2, column=1, value="----------------------------------------------------------------------")
        
        # 计算结果数据
        calc_data = []
        max_settlement = 0
        for point in points:
            # 提取坐标和各种沉降值
            x, y = point.get('x', 0), point.get('y', 0)
            z = point.get('z', 0)
            
            # 提取桩1、桩2和总沉降值
            pile1_settlement = point.get('pile1_settlement', 0) * 1000  # 转为mm
            pile2_settlement = point.get('pile2_settlement', 0) * 1000  # 转为mm
            total_settlement = point.get('settlement_mm', 0)  # 已经是mm
            
            # 更新最大沉降值
            if y < 0:  # 只考虑路基下方的点（y坐标为负值的点）
                max_settlement = max(max_settlement, total_settlement)
            
            calc_data.append([
                point.get('point_id', ''),
                f"({x:.1f}, {y:.1f})",
                f"{pile1_settlement:.2f}",
                f"{pile2_settlement:.2f}",
                f"{total_settlement:.2f}"
            ])
        
        # 创建DataFrame并写入Excel
        start_row = title_row + 3  # 表格开始行
        for i, row_data in enumerate(calc_data):
            for j, value in enumerate(row_data):
                writer.sheets[sheet_name].cell(row=start_row+i, column=j+1, value=value)
        
        # 在表格下方添加最大沉降值信息
        end_row = start_row + len(calc_data) + 1
        bridge_limit = safety_assessment.get('bridge_limit', 150)  # 默认桥梁限值150mm
        writer.sheets[sheet_name].cell(row=end_row, column=1, 
                                    value=f"路基下的计算点总沉降量最大值: {max_settlement:.2f} mm → " + 
                                         (f"超过容许值({bridge_limit}mm)" if max_settlement > bridge_limit else f"未超过容许值({bridge_limit}mm)"))
        
        # 添加公路等级与沉降容许值对比说明
        writer.sheets[sheet_name].cell(row=end_row+2, column=1, 
                                    value="注：根据《公路路基设计规范》JTG D30-2015，不同公路等级的沉降容许值要求不同。")
        
        # 格式调整（可根据需要进一步优化）
        for i in range(1, len(calc_data[0])+1):
            writer.sheets[sheet_name].column_dimensions[chr(64+i)].width = 15
    
    def _write_statistics(self, writer, results):
        """写入统计分析表"""
        stats = results['statistics']
        
        stats_data = [
            ['统计项目', '数值', '单位'],
            ['最大沉降值', f"{stats['max_settlement_mm']:.3f}", 'mm'],
            ['最小沉降值', f"{stats['min_settlement_mm']:.3f}", 'mm'],
            ['平均沉降值', f"{stats['avg_settlement_mm']:.3f}", 'mm'],
            ['沉降值范围', f"{stats['max_settlement_mm'] - stats['min_settlement_mm']:.3f}", 'mm'],
        ]
        
        stats_df = pd.DataFrame(stats_data[1:], columns=stats_data[0])
        stats_df.to_excel(writer, sheet_name='统计分析', index=False)
    
    def _write_safety_assessment(self, writer, results):
        """写入安全评估表"""
        safety = results['safety_assessment']
        
        safety_data = [
            ['评估项目', '结果', '说明'],
            ['安全等级', safety['safety_level'], ''],
            ['最大沉降值', f"{safety['max_settlement_mm']:.3f} mm", ''],
            ['一般限值', f"{safety['general_limit']} mm", ''],
            ['桥梁限值', f"{safety['bridge_limit']} mm", ''],
            ['是否超过一般限值', '是' if safety['exceed_general'] else '否', ''],
            ['是否超过桥梁限值', '是' if safety['exceed_bridge'] else '否', ''],
            ['影响范围面积', f"{safety['influence_area']:.2f} m²", ''],
            ['影响点数量', f"{safety['influence_points_count']}", '个'],
        ]
        
        safety_df = pd.DataFrame(safety_data[1:], columns=safety_data[0])
        safety_df.to_excel(writer, sheet_name='安全评估', index=False, startrow=0)
        
        # 建议事项
        recommendations_data = [['序号', '建议事项']]
        for i, rec in enumerate(safety['recommendations'], 1):
            recommendations_data.append([i, rec])
        
        if len(recommendations_data) > 1:
            rec_df = pd.DataFrame(recommendations_data[1:], columns=recommendations_data[0])
            rec_df.to_excel(writer, sheet_name='安全评估', index=False, 
                          startrow=len(safety_data)+2)
    
    def _write_calculation_diagram(self, writer, results, params):
        """在Excel中嵌入计算简图"""
        try:
            import tempfile
            import os
            from visualization.plotter import ResultPlotter
            
            # 创建计算简图表
            sheet_name = '计算简图'
            writer.sheets[sheet_name] = writer.book.create_sheet(sheet_name)
            
            # 添加标题
            writer.sheets[sheet_name].cell(row=1, column=1, value="计算简图:")
            writer.sheets[sheet_name].cell(row=2, column=1, value="----------------------------------------------------------------------")
            
            # 生成并保存计算简图
            plotter = ResultPlotter()
            fig = plotter.settlement_distribution_plot(results, params)
            
            # 创建临时文件保存图像
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                tmp_path = tmp.name
            
            # 保存图像
            fig.savefig(tmp_path, dpi=300, bbox_inches='tight')
            
            # 在Excel中插入图像
            img = openpyxl.drawing.image.Image(tmp_path)
            img.width = 800  # 图像宽度（像素）
            img.height = 600  # 图像高度（像素）
            writer.sheets[sheet_name].add_image(img, 'A4')  # 在A4单元格位置插入图像
            
            # 删除临时文件
            os.unlink(tmp_path)
            
            return True
        except Exception as e:
            print(f"嵌入计算简图失败: {str(e)}")
            # 在Excel中添加错误信息
            writer.sheets[sheet_name].cell(row=4, column=1, 
                                        value=f"无法嵌入计算简图: {str(e)}")
            return False
    
    def _write_correction_factors(self, writer, results):
        """写入修正系数表"""
        # 检查correction_factors是否存在，如果不存在则创建默认值
        correction = results.get('correction_factors', {
            'length_correction': 1.0,
            'diameter_correction': 1.0,
            'combined_correction': 1.0
        })
        
        # 获取桩的参数（处理双桩情况）
        input_params = results['input_parameters']
        if 'pile1' in input_params:
            # 双桩结构，使用桩1的参数作为主要参数
            pile_length = input_params['pile1']['length']
            pile_diameter = input_params['pile1']['diameter']
        else:
            # 单桩结构
            pile_length = input_params.get('pile_length', 20.0)
            pile_diameter = input_params.get('pile_diameter', 1.0)
        
        correction_data = [
            ['修正系数', '数值', '说明'],
            ['桩长修正系数a', f"{correction.get('length_correction', 1.0):.4f}", 
             f"a = 0.985 - 0.00051 × {pile_length}"],
            ['桩径修正系数b', f"{correction.get('diameter_correction', 1.0):.4f}", 
             f"b = 0.038 × {pile_diameter}² - 0.206 × {pile_diameter} + 1.159"],
            ['综合修正系数', f"{correction.get('combined_correction', 1.0):.4f}", 'a × b'],
        ]
        
        correction_df = pd.DataFrame(correction_data[1:], columns=correction_data[0])
        correction_df.to_excel(writer, sheet_name='修正系数', index=False)
    
    def export_to_csv(self, results, output_dir):
        """
        导出结果到CSV文件
        
        参数:
        results: 计算结果字典
        output_dir: 输出目录
        """
        try:
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 导出计算结果
            points = results['points']
            calc_data = []
            for point in points:
                calc_data.append({
                    '计算点': point['point_id'],
                    'X坐标(m)': point['x'],
                    'Y坐标(m)': point['y'],
                    '深度(m)': point['z'],
                    '距离(m)': point['distance'],
                    '沉降值(mm)': point.get('settlement_mm', 0),
                    '影响系数': point['influence_factor'],
                    '垂直应力(kPa)': point['sigma_z'],
                    '剪应力τxz(kPa)': point['tau_xz'],
                    '剪应力τyz(kPa)': point['tau_yz'],
                    '对应土层': point['soil_properties']['name']
                })
            
            calc_df = pd.DataFrame(calc_data)
            calc_df.to_csv(os.path.join(output_dir, '计算结果.csv'), 
                          index=False, encoding='utf-8-sig')
            
            return True, "CSV导出成功"
            
        except Exception as e:
            return False, f"CSV导出失败：{str(e)}"
    
    def export_to_json(self, results, filename):
        """
        导出结果到JSON文件
        
        参数:
        results: 计算结果字典
        filename: 输出文件名
        """
        try:
            # 添加导出时间戳
            export_data = {
                'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'software_info': {
                    'name': '桥梁跨越工程安全性评估软件',
                    'version': '1.0',
                    'author': '金洪松'
                },
                'results': results
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            
            return True, "JSON导出成功"
            
        except Exception as e:
            return False, f"JSON导出失败：{str(e)}"
    
    def generate_summary_report(self, results):
        """
        生成汇总报告文本
        
        参数:
        results: 计算结果字典
        
        返回:
        report_text: 报告文本
        """
        input_params = results['input_parameters']
        stats = results['statistics']
        safety = results['safety_assessment']
        
        report_lines = [
            "=" * 60,
            "高架桥桩基沉降影响范围计算报告",
            "=" * 60,
            "",
            f"计算时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}",
            f"软件版本：桥梁跨越工程安全性评估软件 1.0",
            f"计算方法：Boussinesq理论 + 参数修正",
            f"技术规范：JTG D30-2015",
            "",
            "一、输入参数",
            "-" * 30,
            f"桩径：{input_params['pile_diameter']} m",
            f"桩长：{input_params['pile_length']} m",
            f"荷载：{input_params['load']} kN",
            f"路线等级：{input_params['road_level']}",
            f"土层数量：{len(input_params['soil_layers'])} 层",
            "",
            "二、计算结果统计",
            "-" * 30,
            f"最大沉降值：{stats['max_settlement_mm']:.3f} mm",
            f"最小沉降值：{stats['min_settlement_mm']:.3f} mm",
            f"平均沉降值：{stats['avg_settlement_mm']:.3f} mm",
            f"计算点数量：{len(results['points'])} 个",
            "",
            "三、安全评估",
            "-" * 30,
            f"安全等级：{safety['safety_level']}",
            f"限值标准：一般限值 {safety['general_limit']} mm，桥梁限值 {safety['bridge_limit']} mm",
            f"超限情况：{'存在超限' if safety['exceed_bridge'] else '未超限'}",
            f"影响范围：{safety['influence_area']:.2f} m²",
            "",
            "四、工程建议",
            "-" * 30,
        ]
        
        for i, rec in enumerate(safety['recommendations'], 1):
            report_lines.append(f"{i}. {rec}")
        
        report_lines.extend([
            "",
            "五、计算依据",
            "-" * 30,
            "1. 《公路路基设计规范》JTG D30-2015",
            "2. Boussinesq弹性理论",
            "3. FLAC3D数值模拟修正",
            "",
            "=" * 60,
            "报告结束",
            "=" * 60
        ])
        
        return "\n".join(report_lines)
    
    def export_summary_report(self, results, filename):
        """
        导出汇总报告到文本文件
        
        参数:
        results: 计算结果字典
        filename: 输出文件名
        """
        try:
            report_text = self.generate_summary_report(results)
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(report_text)
            
            return True, "汇总报告导出成功"
            
        except Exception as e:
            return False, f"汇总报告导出失败：{str(e)}"
    
    def get_export_formats(self):
        """获取支持的导出格式"""
        return {
            'excel': {
                'name': 'Excel工作簿',
                'extension': '.xlsx',
                'description': '包含完整计算结果的Excel文件'
            },
            'csv': {
                'name': 'CSV数据表',
                'extension': '.csv',
                'description': '计算结果数据表格'
            },
            'json': {
                'name': 'JSON数据',
                'extension': '.json',
                'description': '结构化数据格式'
            },
            'txt': {
                'name': '汇总报告',
                'extension': '.txt',
                'description': '简要的文本报告'
            }
        } 